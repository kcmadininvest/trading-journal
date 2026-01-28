from io import BytesIO
from datetime import datetime
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """
    Générateur de rapports Excel pour les statistiques de portefeuille.
    Utilise openpyxl pour créer des fichiers Excel avec graphiques.
    """
    
    def __init__(self, trading_account, stats: Dict[str, Any], config: Dict[str, Any]):
        """
        Initialise le générateur Excel.
        
        Args:
            trading_account: Instance de TradingAccount
            stats: Dictionnaire des statistiques calculées
            config: Configuration de l'export (sections à inclure)
        """
        self.trading_account = trading_account
        self.stats = stats
        self.config = config
        self.wb = Workbook()
        
    def generate(self) -> BytesIO:
        """
        Génère le fichier Excel et retourne un BytesIO.
        
        Returns:
            BytesIO contenant le fichier Excel généré
        """
        if self.wb.active is not None:
            self.wb.remove(self.wb.active)
        
        sections = self.config.get('sections', {})
        
        self._create_summary_sheet()
        
        if 'metrics' in sections:
            self._create_metrics_sheet()
        
        if 'analysis' in sections:
            analysis = sections.get('analysis', [])
            if 'by_strategy' in analysis:
                self._create_strategy_sheet()
            if 'by_instrument' in analysis:
                self._create_instrument_sheet()
            if 'by_timeframe' in analysis:
                self._create_timeframe_sheet()
            if 'by_day_of_week' in analysis:
                self._create_day_of_week_sheet()
        
        if 'charts' in sections:
            charts = sections.get('charts', [])
            if 'equity_curve' in charts:
                self._create_equity_curve_sheet()
            if 'monthly_performance' in charts:
                self._create_monthly_performance_sheet()
        
        trades_list = sections.get('trades_list')
        if trades_list and trades_list != 'none':
            self._create_trades_sheet(trades_list)
        
        excel_file = BytesIO()
        self.wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _create_summary_sheet(self):
        """Crée la feuille de résumé."""
        ws = self.wb.create_sheet("Résumé", 0)
        
        ws['A1'] = "RAPPORT DE PERFORMANCE"
        ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Compte de Trading:"
        ws['B3'] = self.trading_account.name
        ws['A4'] = "Période:"
        period_start = self.config.get('options', {}).get('period_start', 'Début')
        period_end = self.config.get('options', {}).get('period_end', 'Aujourd\'hui')
        ws['B4'] = f"{period_start} - {period_end}"
        ws['A5'] = "Généré le:"
        ws['B5'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        
        for cell in ['A3', 'A4', 'A5']:
            ws[cell].font = Font(bold=True)
        
        general = self.stats.get('general', {})
        performance = self.stats.get('performance', {})
        risk = self.stats.get('risk', {})
        
        row = 7
        ws[f'A{row}'] = "STATISTIQUES PRINCIPALES"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 2
        stats_data = [
            ("Capital Initial", f"{general.get('initial_capital', 0):,.2f} €"),
            ("Capital Actuel", f"{general.get('current_capital', 0):,.2f} €"),
            ("P&L Net", f"{general.get('net_pnl', 0):,.2f} €"),
            ("Rendement", f"{general.get('return_pct', 0):.2f} %"),
            ("", ""),
            ("Nombre de Trades", str(general.get('total_trades', 0))),
            ("Trades Gagnants", str(general.get('winning_trades', 0))),
            ("Trades Perdants", str(general.get('losing_trades', 0))),
            ("Win Rate", f"{general.get('win_rate', 0):.2f} %"),
            ("", ""),
            ("Profit Factor", f"{performance.get('profit_factor', 0):.2f}"),
            ("Expectancy", f"{performance.get('expectancy', 0):.2f} €"),
            ("Ratio R/R Moyen", f"{performance.get('risk_reward_ratio', 0):.2f}"),
            ("", ""),
            ("Drawdown Max", f"{risk.get('max_drawdown', 0):,.2f} €"),
            ("Drawdown Max %", f"{risk.get('max_drawdown_pct', 0):.2f} %"),
            ("Sharpe Ratio", f"{risk.get('sharpe_ratio', 0):.2f}"),
        ]
        
        for label, value in stats_data:
            if label:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                
                if any(keyword in label.lower() for keyword in ['p&l', 'rendement', 'capital actuel']):
                    try:
                        val = float(value.replace('€', '').replace('%', '').replace(',', '').strip())
                        if val > 0:
                            ws[f'B{row}'].font = Font(color="00B050")
                        elif val < 0:
                            ws[f'B{row}'].font = Font(color="FF0000")
                    except:
                        pass
            row += 1
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _create_metrics_sheet(self):
        """Crée la feuille des métriques détaillées."""
        ws = self.wb.create_sheet("Métriques")
        
        ws['A1'] = "MÉTRIQUES DÉTAILLÉES"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E78")
        ws.merge_cells('A1:B1')
        
        general = self.stats.get('general', {})
        performance = self.stats.get('performance', {})
        risk = self.stats.get('risk', {})
        
        row = 3
        sections_data = [
            ("GÉNÉRAL", [
                ("Total Trades", general.get('total_trades', 0)),
                ("Trades Gagnants", general.get('winning_trades', 0)),
                ("Trades Perdants", general.get('losing_trades', 0)),
                ("Trades Breakeven", general.get('breakeven_trades', 0)),
                ("Win Rate", f"{general.get('win_rate', 0):.2f}%"),
                ("P&L Total", f"{general.get('total_pnl', 0):,.2f} €"),
                ("Frais Totaux", f"{general.get('total_fees', 0):,.2f} €"),
                ("P&L Net", f"{general.get('net_pnl', 0):,.2f} €"),
            ]),
            ("PERFORMANCE", [
                ("Gain Moyen", f"{performance.get('average_win', 0):,.2f} €"),
                ("Perte Moyenne", f"{performance.get('average_loss', 0):,.2f} €"),
                ("Plus Gros Gain", f"{performance.get('largest_win', 0):,.2f} €"),
                ("Plus Grosse Perte", f"{performance.get('largest_loss', 0):,.2f} €"),
                ("Profit Factor", f"{performance.get('profit_factor', 0):.2f}"),
                ("Expectancy", f"{performance.get('expectancy', 0):.2f} €"),
                ("Ratio R/R Moyen", f"{performance.get('risk_reward_ratio', 0):.2f}"),
            ]),
            ("RISQUE", [
                ("Drawdown Max", f"{risk.get('max_drawdown', 0):,.2f} €"),
                ("Drawdown Max %", f"{risk.get('max_drawdown_pct', 0):.2f}%"),
                ("Drawdown Actuel", f"{risk.get('current_drawdown', 0):,.2f} €"),
                ("Sharpe Ratio", f"{risk.get('sharpe_ratio', 0):.2f}"),
            ]),
        ]
        
        for section_title, section_data in sections_data:
            ws[f'A{row}'] = section_title
            ws[f'A{row}'].font = Font(size=12, bold=True, color="1F4E78")
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            for label, value in section_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_strategy_sheet(self):
        """Crée la feuille d'analyse par stratégie."""
        ws = self.wb.create_sheet("Par Stratégie")
        
        headers = ["Stratégie", "Total Trades", "Trades Gagnants", "Win Rate %", "P&L Total", "P&L Moyen"]
        self._write_table_headers(ws, headers)
        
        strategy_stats = self.stats.get('by_strategy', [])
        
        row = 2
        for stat in strategy_stats:
            ws[f'A{row}'] = stat.get('strategy', '')
            ws[f'B{row}'] = stat.get('total_trades', 0)
            ws[f'C{row}'] = stat.get('winning_trades', 0)
            ws[f'D{row}'] = stat.get('win_rate', 0)
            ws[f'E{row}'] = stat.get('total_pnl', 0)
            ws[f'F{row}'] = stat.get('avg_pnl', 0)
            
            if stat.get('total_pnl', 0) > 0:
                ws[f'E{row}'].font = Font(color="00B050")
            elif stat.get('total_pnl', 0) < 0:
                ws[f'E{row}'].font = Font(color="FF0000")
            
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
    
    def _create_instrument_sheet(self):
        """Crée la feuille d'analyse par instrument."""
        ws = self.wb.create_sheet("Par Instrument")
        
        headers = ["Instrument", "Total Trades", "Trades Gagnants", "Win Rate %", "P&L Total", "P&L Moyen"]
        self._write_table_headers(ws, headers)
        
        instrument_stats = self.stats.get('by_instrument', [])
        
        row = 2
        for stat in instrument_stats:
            ws[f'A{row}'] = stat.get('instrument', '')
            ws[f'B{row}'] = stat.get('total_trades', 0)
            ws[f'C{row}'] = stat.get('winning_trades', 0)
            ws[f'D{row}'] = stat.get('win_rate', 0)
            ws[f'E{row}'] = stat.get('total_pnl', 0)
            ws[f'F{row}'] = stat.get('avg_pnl', 0)
            
            if stat.get('total_pnl', 0) > 0:
                ws[f'E{row}'].font = Font(color="00B050")
            elif stat.get('total_pnl', 0) < 0:
                ws[f'E{row}'].font = Font(color="FF0000")
            
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
    
    def _create_timeframe_sheet(self):
        """Crée la feuille d'analyse par timeframe."""
        ws = self.wb.create_sheet("Par Timeframe")
        
        headers = ["Timeframe", "Total Trades", "Trades Gagnants", "Win Rate %", "P&L Total"]
        self._write_table_headers(ws, headers)
        
        timeframe_stats = self.stats.get('by_timeframe', [])
        
        row = 2
        for stat in timeframe_stats:
            ws[f'A{row}'] = stat.get('timeframe', '')
            ws[f'B{row}'] = stat.get('total_trades', 0)
            ws[f'C{row}'] = stat.get('winning_trades', 0)
            ws[f'D{row}'] = stat.get('win_rate', 0)
            ws[f'E{row}'] = stat.get('total_pnl', 0)
            
            if stat.get('total_pnl', 0) > 0:
                ws[f'E{row}'].font = Font(color="00B050")
            elif stat.get('total_pnl', 0) < 0:
                ws[f'E{row}'].font = Font(color="FF0000")
            
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def _create_day_of_week_sheet(self):
        """Crée la feuille d'analyse par jour de la semaine."""
        ws = self.wb.create_sheet("Par Jour")
        
        headers = ["Jour", "Total Trades", "Trades Gagnants", "Win Rate %", "P&L Total"]
        self._write_table_headers(ws, headers)
        
        day_stats = self.stats.get('by_day_of_week', [])
        
        row = 2
        for stat in day_stats:
            ws[f'A{row}'] = stat.get('day', '')
            ws[f'B{row}'] = stat.get('total_trades', 0)
            ws[f'C{row}'] = stat.get('winning_trades', 0)
            ws[f'D{row}'] = stat.get('win_rate', 0)
            ws[f'E{row}'] = stat.get('total_pnl', 0)
            
            if stat.get('total_pnl', 0) > 0:
                ws[f'E{row}'].font = Font(color="00B050")
            elif stat.get('total_pnl', 0) < 0:
                ws[f'E{row}'].font = Font(color="FF0000")
            
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def _create_equity_curve_sheet(self):
        """Crée la feuille avec la courbe d'équité."""
        ws = self.wb.create_sheet("Courbe d'Équité")
        
        equity_data = self.stats.get('equity_curve', [])
        
        if not equity_data:
            return
        
        ws['A1'] = "Date"
        ws['B1'] = "Capital"
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)
        
        row = 2
        for point in equity_data:
            ws[f'A{row}'] = point.get('date', '')
            ws[f'B{row}'] = point.get('balance', 0)
            row += 1
        
        chart = LineChart()
        chart.title = "Courbe d'Équité"
        chart.style = 13
        chart.y_axis.title = "Capital (€)"
        chart.x_axis.title = "Date"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        
        ws.add_chart(chart, "D2")
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_monthly_performance_sheet(self):
        """Crée la feuille avec la performance mensuelle."""
        ws = self.wb.create_sheet("Performance Mensuelle")
        
        monthly_data = self.stats.get('monthly_performance', [])
        
        if not monthly_data:
            return
        
        ws['A1'] = "Mois"
        ws['B1'] = "P&L"
        ws['C1'] = "Trades"
        ws['D1'] = "Win Rate %"
        
        for cell in ['A1', 'B1', 'C1', 'D1']:
            ws[cell].font = Font(bold=True)
        
        row = 2
        for month in monthly_data:
            ws[f'A{row}'] = month.get('month', '')
            ws[f'B{row}'] = month.get('total_pnl', 0)
            ws[f'C{row}'] = month.get('total_trades', 0)
            ws[f'D{row}'] = month.get('win_rate', 0)
            
            if month.get('total_pnl', 0) > 0:
                ws[f'B{row}'].font = Font(color="00B050")
            elif month.get('total_pnl', 0) < 0:
                ws[f'B{row}'].font = Font(color="FF0000")
            
            row += 1
        
        chart = BarChart()
        chart.title = "Performance Mensuelle"
        chart.style = 10
        chart.y_axis.title = "P&L (€)"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "F2")
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
    
    def _create_trades_sheet(self, trades_type: str):
        """Crée la feuille avec la liste des trades."""
        ws = self.wb.create_sheet("Trades")
        
        top_trades = self.stats.get('top_trades', {})
        all_trades = self.stats.get('all_trades', [])
        
        headers = ["Date", "Instrument", "Stratégie", "Direction", "P&L", "P&L %", "Notes"]
        self._write_table_headers(ws, headers)
        
        trades_to_show = []
        
        if trades_type == 'top_10_best_worst':
            trades_to_show = top_trades.get('best', [])[:10] + top_trades.get('worst', [])[:10]
        elif trades_type == 'top_10_best':
            trades_to_show = top_trades.get('best', [])[:10]
        elif trades_type == 'top_10_worst':
            trades_to_show = top_trades.get('worst', [])[:10]
        elif trades_type == 'all':
            trades_to_show = all_trades
        
        row = 2
        for trade in trades_to_show:
            date_value = trade.get('date', '')
            if hasattr(date_value, 'strftime'):
                date_value = date_value.strftime('%d/%m/%Y')
            
            ws[f'A{row}'] = date_value
            ws[f'B{row}'] = trade.get('instrument', '')
            ws[f'C{row}'] = trade.get('strategy', '')
            ws[f'D{row}'] = trade.get('direction', '')
            ws[f'E{row}'] = trade.get('pnl', 0)
            ws[f'F{row}'] = trade.get('pnl_pct', 0)
            ws[f'G{row}'] = trade.get('notes', '')
            
            if trade.get('pnl', 0) > 0:
                ws[f'E{row}'].font = Font(color="00B050")
            elif trade.get('pnl', 0) < 0:
                ws[f'E{row}'].font = Font(color="FF0000")
            
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['G'].width = 30
    
    def _write_table_headers(self, ws, headers):
        """Écrit les en-têtes d'un tableau."""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
